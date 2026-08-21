# -*- coding: utf-8 -*-
"""
run_eval.py — ตัวรันชุดทดสอบ 100 ลิงก์ แล้วออกไฟล์ Excel รูปแบบเดิมให้อัตโนมัติ
============================================================================

วิธีใช้ (วางไฟล์นี้กับ testset_100.json ไว้ในโฟลเดอร์ backend แล้วสั่ง):

    python run_eval.py                  # ตรวจเต็มทุกชั้น (ช้า แต่ได้ผลจริง)
    python run_eval.py --fast           # ตรวจแค่ชั้น 1-2 (เร็ว ไม่ต้องต่อเน็ตมาก)
    python run_eval.py --workers 8      # เร่งความเร็วด้วยการตรวจพร้อมกันหลายลิงก์
    python run_eval.py --limit 20       # ลองแค่ 20 ลิงก์แรกก่อน (ไว้เทสต์ว่ารันได้)
    python run_eval.py --out ผลรอบ2.xlsx

ต้องมี:  pandas, openpyxl   (ถ้ายังไม่มี:  pip install pandas openpyxl)

*** คำเตือน: ลิงก์กลุ่ม phish_fresh ในชุดทดสอบเป็นเว็บหลอกที่ยังทำงานอยู่จริง
    สคริปต์นี้แค่ "ส่งให้ระบบตรวจ" ไม่ได้เปิดหน้าเว็บให้คุณดู อย่าเอาลิงก์ไปเปิดเองเด็ดขาด ***
"""
import argparse
import json
import os
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

FONT = "Tahoma"
COLS = ["ลำดับ", "ลิงก์ (url)", "ความจริง (label)", "กลุ่ม (stratum)",
        "แหล่งที่มา (source)", "ผลที่ระบบตอบ", "คะแนน", "สัญญาณที่เจอ", "หมายเหตุ"]


def load_testset(path: str) -> list:
    if not os.path.exists(path):
        sys.exit(f"หาไฟล์ชุดทดสอบไม่เจอ: {path}\n"
                 f"เอา testset_100.json มาวางไว้โฟลเดอร์เดียวกับสคริปต์นี้")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def show_blocklist_status():
    """บอกให้ชัดว่าชั้นที่ 1 (บัญชีดำ สกมช.) ใช้งานได้จริงหรือเปล่าตอนรัน
    ถ้าโหลดไม่ได้ ผลที่ออกมาจะต่ำกว่าความจริง ต้องรู้ก่อนอ่านตัวเลข"""
    try:
        from analyzer.blacklist_api import load_blocklist
        n = load_blocklist()
        if n:
            print(f"  ชั้น 1 บัญชีดำ สกมช. : พร้อมใช้ ({n:,} โดเมน)")
        else:
            print("  ชั้น 1 บัญชีดำ สกมช. : *** โหลดไม่ได้ *** — ผลที่ได้จะต่ำกว่าความจริง")
        return n
    except Exception as e:
        print(f"  ชั้น 1 บัญชีดำ สกมช. : เรียกใช้ไม่ได้ ({type(e).__name__})")
        return 0


def scan_one(item: dict, run_deep: bool) -> dict:
    from analyzer.scanner import scan
    out = dict(item)
    try:
        res = scan(item["url"], run_deep=run_deep)
        if not res.get("ok"):
            out.update(color="error", score=0, sig=res.get("error", "")[:60])
        else:
            out.update(color=res["verdict"]["color"],
                       score=res.get("score", 0),
                       sig="|".join(s["id"] for s in res.get("reasons", [])))
    except Exception as e:                      # ลิงก์เดียวพังต้องไม่ทำให้ทั้งรอบพัง
        out.update(color="error", score=0, sig=f"{type(e).__name__}: {e}"[:60])
    return out


def is_correct(row: dict) -> bool:
    return (row["color"] != "red") if row["label"] == "safe" else (row["color"] == "red")


def write_excel(rows: list, path: str, run_deep: bool, blocklist_n: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "testset_100"
    hfill = PatternFill("solid", fgColor="44546A")
    hfont = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    warn = ws.cell(1, 1, "คำเตือน: ลิงก์กลุ่ม phish_fresh เป็นเว็บหลอกที่ยังทำงานอยู่จริง "
                         "— ห้ามเปิด ใช้สำหรับทดสอบระบบเท่านั้น")
    warn.font = Font(name=FONT, bold=True, size=10, color="9C0006")
    warn.fill = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    for j, c in enumerate(COLS, 1):
        cell = ws.cell(2, j, c)
        cell.fill, cell.font, cell.border = hfill, hfont, bd
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CF = {"green": "C6EFCE", "yellow": "FFEB9C", "red": "FFC7CE", "error": "D9D9D9"}
    CT = {"green": "006100", "yellow": "9C5700", "red": "9C0006", "error": "3F3F3F"}
    for i, r in enumerate(rows):
        rr = 3 + i
        vals = [i + 1, r["url"], r["label"], r["stratum"], r["source"],
                r["color"], r["score"], r["sig"], r["note"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, v)
            c.font = Font(name=FONT, size=10)
            c.border = bd
            c.alignment = Alignment(vertical="top", wrap_text=(j in (2, 8, 9)),
                                    horizontal="center" if j in (1, 7) else "left")
        vc = ws.cell(rr, 6)
        vc.fill = PatternFill("solid", fgColor=CF.get(r["color"], "D9D9D9"))
        vc.font = Font(name=FONT, size=10, bold=True, color=CT.get(r["color"], "3F3F3F"))
        vc.alignment = Alignment(horizontal="center", vertical="top")
        if not is_correct(r):                    # ทำเครื่องหมายแถวที่ระบบตอบผิด
            ws.cell(rr, 1).fill = PatternFill("solid", fgColor="FFF2CC")

    for j, w in enumerate([7, 62, 14, 16, 20, 13, 8, 30, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:I{2 + len(rows)}"

    # ---------------- ชีตสรุป (ใช้สูตรจริง แก้ข้อมูลแล้วตัวเลขขยับตาม) --------------
    D, FIRST, LAST = "testset_100", 3, 2 + len(rows)
    s = wb.create_sheet("สรุปผล")
    T = Font(name=FONT, bold=True, size=13)
    H = Font(name=FONT, bold=True, size=10, color="FFFFFF")
    B = Font(name=FONT, size=10)
    r = 1
    s.cell(r, 1, "สรุปผลการทดสอบ 100 ลิงก์").font = T
    r += 1
    mode = "ตรวจเต็มทุกชั้น (run_deep=True)" if run_deep else "ตรวจแค่ชั้น 1-2 (run_deep=False)"
    bl = f"บัญชีดำ สกมช. {blocklist_n:,} โดเมน" if blocklist_n else "บัญชีดำ สกมช. โหลดไม่ได้"
    s.cell(r, 1, f"โหมด: {mode}  |  {bl}  |  วันที่รัน: "
                 f"{time.strftime('%d/%m/%Y %H:%M')}").font = Font(name=FONT, size=9,
                                                                   italic=True, color="595959")
    r += 2

    s.cell(r, 1, "1) ตารางความสับสน (Confusion Matrix)").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for j, h in enumerate(["", "ระบบตอบ แดง", "ระบบตอบ ไม่แดง", "รวม"], 1):
        c = s.cell(r, j, h); c.font = H; c.fill = hfill; c.border = bd
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    s.cell(r, 1, "ความจริง: อันตราย").font = B
    s.cell(r, 2, f'=COUNTIFS({D}!$C${FIRST}:$C${LAST},"dangerous",{D}!$F${FIRST}:$F${LAST},"red")')
    s.cell(r, 3, f'=COUNTIFS({D}!$C${FIRST}:$C${LAST},"dangerous")-B{r}')
    s.cell(r, 4, f"=B{r}+C{r}")
    tp = r; r += 1
    s.cell(r, 1, "ความจริง: ปลอดภัย").font = B
    s.cell(r, 2, f'=COUNTIFS({D}!$C${FIRST}:$C${LAST},"safe",{D}!$F${FIRST}:$F${LAST},"red")')
    s.cell(r, 3, f'=COUNTIFS({D}!$C${FIRST}:$C${LAST},"safe")-B{r}')
    s.cell(r, 4, f"=B{r}+C{r}")
    fp = r
    for rr2 in (tp - 1, tp, fp):
        for j in range(1, 5):
            s.cell(rr2, j).border = bd
    r += 2

    s.cell(r, 1, "2) ตัวชี้วัด").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for j, h in enumerate(["ตัวชี้วัด", "ค่า", "ความหมาย"], 1):
        c = s.cell(r, j, h); c.font = H; c.fill = hfill; c.border = bd
        c.alignment = Alignment(horizontal="center")
    r += 1
    metrics = [
        ("จับลิงก์หลอกได้ (Recall)", f"=B{tp}/D{tp}", "ลิงก์อันตรายทั้งหมด เตือนแดงได้กี่ %"),
        ("เตือนผิด (False Positive rate)", f"=B{fp}/D{fp}", "เว็บจริงโดนเตือนแดงผิดกี่ % (ยิ่งต่ำยิ่งดี)"),
        ("ความแม่นของคำเตือน (Precision)",
         f"=IF((B{tp}+B{fp})=0,0,B{tp}/(B{tp}+B{fp}))", "เวลาขึ้นแดง ถูกจริงกี่ %"),
        ("ความถูกต้องรวม (Accuracy)", f"=(B{tp}+C{fp})/(D{tp}+D{fp})", "ตอบถูกกี่ % จากทั้งหมด"),
        ('ลิงก์หลอกที่ระบบตอบ "เขียว"',
         f'=COUNTIFS({D}!$C${FIRST}:$C${LAST},"dangerous",{D}!$F${FIRST}:$F${LAST},"green")/D{tp}',
         "อันตรายที่สุด: ระบบรับรองว่าปลอดภัยทั้งที่เป็นเว็บหลอก"),
    ]
    for name, f, desc in metrics:
        s.cell(r, 1, name).font = B
        c = s.cell(r, 2, f); c.font = Font(name=FONT, size=10, bold=True)
        c.number_format = "0.0%"; c.alignment = Alignment(horizontal="center")
        s.cell(r, 3, desc).font = B
        for j in (1, 2, 3):
            s.cell(r, j).border = bd
        r += 1
    r += 1

    s.cell(r, 1, "3) แยกตามกลุ่มตัวอย่าง").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for j, h in enumerate(["กลุ่ม", "ความจริง", "จำนวน", "ตอบถูก", "คิดเป็น %"], 1):
        c = s.cell(r, j, h); c.font = H; c.fill = hfill; c.border = bd
        c.alignment = Alignment(horizontal="center")
    r += 1
    seen = []
    for x in rows:
        key = (x["stratum"], x["label"])
        if key not in seen:
            seen.append(key)
    for st, lab in seen:
        s.cell(r, 1, st).font = B
        s.cell(r, 2, "ปลอดภัย" if lab == "safe" else "อันตราย").font = B
        s.cell(r, 3, f'=COUNTIFS({D}!$D${FIRST}:$D${LAST},A{r})')
        if lab == "safe":
            s.cell(r, 4, f'=C{r}-COUNTIFS({D}!$D${FIRST}:$D${LAST},A{r},'
                         f'{D}!$F${FIRST}:$F${LAST},"red")')
        else:
            s.cell(r, 4, f'=COUNTIFS({D}!$D${FIRST}:$D${LAST},A{r},'
                         f'{D}!$F${FIRST}:$F${LAST},"red")')
        c = s.cell(r, 5, f"=IF(C{r}=0,0,D{r}/C{r})"); c.number_format = "0.0%"
        for j in range(1, 6):
            s.cell(r, j).border = bd
            s.cell(r, j).alignment = Alignment(horizontal="center" if j >= 3 else "left")
        r += 1
    r += 1

    s.cell(r, 1, "หมายเหตุ").font = Font(name=FONT, bold=True, size=11)
    r += 1
    notes = ["แถวที่เลขลำดับมีพื้นสีเหลือง = แถวที่ระบบตอบผิด",
             "ตอบถูก หมายถึง ลิงก์อันตราย->แดง / เว็บจริง->ไม่แดง (เหลืองถือว่าถูก)"]
    if not run_deep:
        notes.append("รอบนี้รันแบบ --fast จึงยังไม่ได้ตรวจชั้น 3-4 ตัวเลขที่ได้เป็นพื้นล่างของระบบ")
    if not blocklist_n:
        notes.append("รอบนี้บัญชีดำ สกมช. โหลดไม่ได้ ผลจึงต่ำกว่าความจริง ควรรันใหม่ตอนเน็ตปกติ")
    for line in notes:
        c = s.cell(r, 1, line); c.font = Font(name=FONT, size=9)
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    for j, w in enumerate([34, 16, 12, 12, 14], 1):
        s.column_dimensions[get_column_letter(j)].width = w

    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="รันชุดทดสอบ 100 ลิงก์")
    ap.add_argument("--fast", action="store_true",
                    help="ตรวจแค่ชั้น 1-2 (ไม่ตามปลายทาง/ไม่อ่านเนื้อหาเว็บ)")
    ap.add_argument("--workers", type=int, default=4, help="ตรวจพร้อมกันกี่ลิงก์ (ค่าเริ่มต้น 4)")
    ap.add_argument("--limit", type=int, default=0, help="ทดลองแค่ N ลิงก์แรก")
    ap.add_argument("--testset", default=os.path.join(HERE, "testset_100.json"))
    ap.add_argument("--out", default=os.path.join(HERE, "ผลทดสอบ100ลิงก์.xlsx"))
    args = ap.parse_args()

    run_deep = not args.fast
    items = load_testset(args.testset)
    if args.limit:
        items = items[:args.limit]

    print("=" * 70)
    print(f"ชุดทดสอบ {len(items)} ลิงก์ | โหมด: "
          f"{'ตรวจเต็มทุกชั้น' if run_deep else 'เร็ว (ชั้น 1-2)'} | ตรวจพร้อมกัน {args.workers} ลิงก์")
    print("=" * 70)
    blocklist_n = show_blocklist_status()
    if run_deep:
        print("  ชั้น 3-4 ต้องต่อเน็ตทีละลิงก์ อาจใช้เวลาหลายนาที รอได้เลย\n")

    t0 = time.perf_counter()
    rows = [None] * len(items)
    done = 0
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        futs = {pool.submit(scan_one, it, run_deep): i for i, it in enumerate(items)}
        for fut in futs:
            pass
        for fut, i in futs.items():
            rows[i] = fut.result()
            done += 1
            mark = "." if is_correct(rows[i]) else "x"
            print(mark, end="", flush=True)
            if done % 50 == 0:
                print(f"  {done}/{len(items)}")
    elapsed = time.perf_counter() - t0
    print(f"\n\nรันเสร็จใน {elapsed:.1f} วินาที ({elapsed/len(items):.2f} วิ/ลิงก์)")

    safe_rows = [r for r in rows if r["label"] == "safe"]
    dang_rows = [r for r in rows if r["label"] == "dangerous"]
    tp = sum(1 for r in dang_rows if r["color"] == "red")
    fp = sum(1 for r in safe_rows if r["color"] == "red")
    green_phish = sum(1 for r in dang_rows if r["color"] == "green")
    errs = sum(1 for r in rows if r["color"] == "error")

    print("-" * 70)
    if dang_rows:
        print(f"จับลิงก์หลอกได้ (Recall)  : {tp}/{len(dang_rows)} = {tp/len(dang_rows)*100:.1f}%")
    if safe_rows:
        print(f"เตือนผิดกับเว็บจริง (FP)  : {fp}/{len(safe_rows)} = {fp/len(safe_rows)*100:.1f}%")
    print(f"ถูกต้องรวม (Accuracy)     : {sum(1 for r in rows if is_correct(r))}/{len(rows)}")
    if green_phish:
        print(f"*** ลิงก์หลอกที่ได้ 'เขียว' : {green_phish} อัน (ต้องแก้ก่อนอย่างอื่น) ***")
    if errs:
        print(f"ตรวจไม่สำเร็จ (error)     : {errs} อัน")
    print(f"สีที่ได้ | ลิงก์หลอก: {dict(Counter(r['color'] for r in dang_rows))}")
    print(f"         | เว็บจริง : {dict(Counter(r['color'] for r in safe_rows))}")

    try:
        write_excel(rows, args.out, run_deep, blocklist_n)
        print(f"\nบันทึกไฟล์แล้ว: {args.out}")
    except ImportError:
        json_out = os.path.splitext(args.out)[0] + ".json"
        json.dump(rows, open(json_out, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"\nยังไม่ได้ติดตั้ง openpyxl จึงบันทึกเป็น {json_out} แทน"
              f"\n(ติดตั้งด้วย: pip install openpyxl แล้วรันใหม่จะได้ไฟล์ Excel)")


if __name__ == "__main__":
    main()
